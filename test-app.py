import unittest
import os
import csv
from openpyxl import load_workbook
from app import extrahuj_vysledky


from app import extrahuj_vysledky

class TestExtrakce(unittest.TestCase):
    def setUp(self):
        # Vzorek HTML podobný Google výstupu (zjednodušený)
        self.test_html = '''
        <html>
            <body>
                <div class="tF2Cxc">
                    <a href="https://example.com">
                        <h3>Příklad titulku</h3>
                    </a>
                </div>
            </body>
        </html>
        '''

    def test_extrahuj_vysledky(self):
        dotaz = "python programming"
        vysledky = extrahuj_vysledky("pizza")

        self.assertIsInstance(vysledky, list)
        self.assertGreater(len(vysledky), 0)
        for vysledek in vysledky:
            self.assertIn("titulek", vysledek)
            self.assertIn("url", vysledek)
            self.assertIsInstance(vysledek["titulek"], str)
            self.assertIsInstance(vysledek["url"], str)


if __name__ == "__main__":
    unittest.main()

def test_csv_export(self):
    vysledky = extrahuj_vysledky("test", html=self.test_html)
    filename = "test_output.csv"

    # Zápis do souboru
    with open(filename, "w", newline='', encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["titulek", "url"])
        writer.writeheader()
        for radek in vysledky:
            writer.writerow(radek)

    # Kontrola, že soubor existuje a má obsah
    self.assertTrue(os.path.exists(filename))

    with open(filename, "r", encoding="utf-8") as f:
        lines = f.readlines()
        self.assertGreater(len(lines), 1)  # alespoň hlavička + 1 řádek

    os.remove(filename)

def test_excel_export(self):
    vysledky = extrahuj_vysledky("test", html=self.test_html)
    filename = "test_output.xlsx"

    # Zápis do souboru
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Titulek", "URL"])
    for radek in vysledky:
        ws.append([radek["titulek"], radek["url"]])
    wb.save(filename)

    # Kontrola existence a obsahu
    self.assertTrue(os.path.exists(filename))
    wb2 = load_workbook(filename)
    sheet = wb2.active
    self.assertEqual(sheet.cell(row=2, column=1).value, "Příklad titulku")

    os.remove(filename)

